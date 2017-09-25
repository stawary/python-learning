import requests
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook
import time
wb = Workbook()
ws = wb.active
ws.title = "职位信息"

def getHTMLText(url):
	try:
		head={'User-Agent': 'Mozilla/5.0'}
		r=requests.get(url,headers=head,timeout=30)
		r.raise_for_status()
		r.encoding = 'UTF-8'
		return r.text
	except:
		return ""

def fillBossinfo(info,html):
    soup=BeautifulSoup(html,"html.parser")
    info_list = []
    info = []
    info2 = []
    info3 =[]
    '''
    以下在抓取职位、公司和薪水时，由于不容易一次抓取，分开抓取后又整合。
    '''
    for i in soup.find_all('td',class_="zwmc"):
        p = re.compile(r'target=\"_blank\">.*?</a>',re.S)
        pos = p.findall(str(i))[0][16:-4]  #职位
        info.append(pos)
    for j in soup.find_all('td',class_="gsmc"):
        com = j.string                    #公司
        info2.append(com)
    for k in soup.find_all('td',class_="zwyx"):
        sal = k.string                   #薪资
        info3.append(sal) 
    for row in range(60):
        info_list.extend([info[row],info2[row],info3[row]])
    print(info_list)
    return info_list


def main():
    page = input('请输入要查询的网页数:')
    page2 = int(page)

    binfo_result = []
    for num in range(page2):
        binfo = []
        url = 'http://sou.zhaopin.com/jobs/searchresult.ashx?in=210500&jl=上海&p='+str(num+1)
        html = getHTMLText(url)
        binfo = fillBossinfo(binfo,html) 
        binfo_result.extend(binfo)
        time.sleep(5) #延时5秒再爬取
    ws['A1']='职位'
    ws['B1']='公司'
    ws['C1']='薪水(元)'
    for num in range(60*page2):
        
        numA='A%s'%(num+2)
        numB='B%s'%(num+2)
        numC='C%s'%(num+2)
        ws[str(numA)]=binfo_result[num*3]
        ws[str(numB)]=binfo_result[num*3+1]
        ws[str(numC)]=binfo_result[num*3+2]
    wb.save('职位信息.xlsx')

main()
