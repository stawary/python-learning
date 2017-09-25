import requests
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "python职位信息"
def getHTMLText(url):
    try:
        head={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.101 Safari/537.36 QIHU 360SE'}
        r = requests.get(url, timeout=30,headers=head)
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        return r.text
    except:
        return ""

def fillBossinfo(binfo,html):
	soup=BeautifulSoup(html,"html.parser")
	binfo_list = []
	for b in soup.find_all("div",class_="job-primary"):
		binfo = []
		d = re.compile(r'<h3 class=\"name\">.*?<')
		pos = d.findall(str(b))[0][17:-1]#正则,需将原tag变为str再检索

		sal = b.span.string

		com = b.find('div',class_="info-comapny")
		e = re.compile(r'<h3 class.*?<')
		com2 = e.findall(str(com))[0][17:-1]#注意findall后结果[0],为第一个结果。
		binfo.extend([com2,pos,sal])
		binfo_list.extend(binfo)
	return binfo_list

def main():
    page = input('请输入要查询的网页数:')
    binfo_result = []
    binfo_result2 = []
    page2=int(page)
    for j in range(page2):
        url='http://www.zhipin.com/c101020100/h_101020100/?query=python&page='+str(j+1)
        binfo = []
        html = getHTMLText(url)
        fillBossinfo(binfo, html)
        binfo_result = fillBossinfo(binfo, html)
        binfo_result2.extend(binfo_result)
    ws['A1'] = '公司'
    ws['B1'] = '职位名称'
    ws['C1'] = '薪水'
    for num in range(15*page2):
        numA = 'A%s'%(num+2)
        numB = 'B%s'%(num+2)
        numC = 'C%s'%(num+2)
        ws[str(numA)] = binfo_result2[num*3]
        ws[str(numB)] = binfo_result2[num*3+1]
        ws[str(numC)] = binfo_result2[num*3+2]
    wb.save('Boss直聘职位信息.xlsx')

if __name__ == '__main__':
	main()
